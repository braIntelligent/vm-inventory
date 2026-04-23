"""
excel_writer.py
Genera el archivo Excel final a partir de una lista de Resource.
"""

import logging
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

COLUMNS = [
    # ── IDENTIDAD ─────────────────────────────────────────────────────────────
    ("NOMBRE",                "nombre"),
    ("ESTADO",                "estado"),
    ("TIPO RECURSO",          "tipo_recurso"),
    ("FUENTE",                "fuente"),
    ("PROYECTO",              "proyecto"),
    # ── UBICACION ─────────────────────────────────────────────────────────────
    ("REGION",                "region"),
    ("ZONA",                  "zona"),
    ("VPC",                   "vpc"),
    ("SUBRED",                "subred"),
    # ── COMPUTO (VMs) ─────────────────────────────────────────────────────────
    ("TIPO MAQUINA",          "tipo_maquina"),
    ("VCPUS",                 "vcpus"),
    ("RAM (GB)",              "ram_gb"),
    # ── ALMACENAMIENTO ────────────────────────────────────────────────────────
    ("DISCO (GB)",            "disco_gb"),
    ("TIPO DISCO",            "tipo_disco"),
    # ── SISTEMA ───────────────────────────────────────────────────────────────
    ("SISTEMA OPERATIVO",     "sistema_operativo"),
    # ── RED ───────────────────────────────────────────────────────────────────
    ("IP INTERNA",            "ip_interna"),
    ("IP EXTERNA",            "ip_externa"),
    # ── TIEMPO ────────────────────────────────────────────────────────────────
    ("FECHA CREACION",        "fecha_creacion"),
    ("ULTIMO ENCENDIDO",      "ultimo_encendido"),
    ("ULTIMO APAGADO",        "ultimo_apagado"),
    ("PROGRAMA DE ENCENDIDO", "programa_encendido"),
    ("ULTIMA ACTUALIZACION",  "ultima_actualizacion"),
    # ── BASE DE DATOS (Cloud SQL — vacias para VMs) ───────────────────────────
    ("MOTOR",                 "motor"),
    ("VERSION DB",            "version"),
    ("TIER",                  "tier"),
    ("ALTA DISPONIBILIDAD",   "alta_disponibilidad"),
    ("REPLICA LECTURA",       "replica_lectura"),
    ("BACKUP AUTOMATICO",     "backup_automatico"),
    ("VENTANA BACKUP",        "ventana_backup"),
    # ── SEGURIDAD / GESTION (manuales) ────────────────────────────────────────
    ("AMBIENTE",              "ambiente"),
    ("CRITICIDAD",            "criticidad"),
    ("CATEGORIA",             "categoria"),
    ("PROVEEDOR RESPONSABLE", "proveedor_responsable"),
    ("PAM",                   "pam"),
    ("VERSION CYLANCE",       "version_cylance"),
    ("VERSION FOCUS",         "version_focus"),
    ("GRUPO TENABLE",         "grupo_tenable"),
    ("DESCRIPCION",           "descripcion"),
]

WIDTHS = {
    "NOMBRE":28, "ESTADO":13, "TIPO RECURSO":18, "FUENTE":10, "PROYECTO":32,
    "REGION":24, "ZONA":26, "VPC":28, "SUBRED":28,
    "TIPO MAQUINA":22, "VCPUS":8, "RAM (GB)":10,
    "DISCO (GB)":11, "TIPO DISCO":16,
    "SISTEMA OPERATIVO":34,
    "IP INTERNA":16, "IP EXTERNA":16,
    "FECHA CREACION":16, "ULTIMO ENCENDIDO":16, "ULTIMO APAGADO":16,
    "PROGRAMA DE ENCENDIDO":28, "ULTIMA ACTUALIZACION":20,
    "MOTOR":12, "VERSION DB":24, "TIER":22,
    "ALTA DISPONIBILIDAD":20, "REPLICA LECTURA":16,
    "BACKUP AUTOMATICO":18, "VENTANA BACKUP":16,
    "AMBIENTE":14, "CRITICIDAD":12, "CATEGORIA":15,
    "PROVEEDOR RESPONSABLE":22, "PAM":8,
    "VERSION CYLANCE":18, "VERSION FOCUS":16,
    "GRUPO TENABLE":18, "DESCRIPCION":35,
}

ROW_COLORS = {
    "ENCENDIDA":      "FFD6F5D6",
    "APAGADA":        "FFFFD6D6",
    "INICIANDO":      "FFFFF3CC",
    "SUSPENDIDA":     "FFFFF3CC",
    "ELIMINADA":      "FFE0E0E0",
    "ACTIVO":         "FFD6F5D6",
    "ENCENDIDO":      "FFD6F5D6",
    "HABILITADA":     "FFD6F5D6",
    "ERROR":          "FFFFD6D6",
    "DEGRADADO":      "FFFFD6D6",
}

HEADER_COLOR = "FF375F7C"
HEADERS      = [h for h, _ in COLUMNS]
KEYS         = [k for _, k in COLUMNS]


def _border():
    thin = Side(style="thin", color="FFB0B0B0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


GKE_COLUMNS = [
    ("NOMBRE CLUSTER",        "nombre"),
    ("ESTADO",                "estado"),
    ("FUENTE",                "fuente"),
    ("PROYECTO",              "proyecto"),
    ("REGION",                "region"),
    ("ZONA",                  "zona"),
    ("VPC",                   "vpc"),
    ("SUBRED",                "subred"),
    ("VERSION KUBERNETES",    "version_kubernetes"),
    ("VERSION NODOS",         "version_nodos"),
    ("AUTOPILOT",             "autopilot"),
    ("TOTAL NODOS",           "total_nodos"),
    ("NODOS MIN",             "nodos_min"),
    ("NODOS MAX",             "nodos_max"),
    ("TIPO MAQUINA NODOS",    "tipo_maquina"),
    ("NODE POOLS",            "node_pools"),
    ("RELEASE CHANNEL",       "release_channel"),
    ("WORKLOAD IDENTITY",     "workload_identity"),
    ("IP CLUSTER (CIDR)",     "ip_interna"),
    ("ENDPOINT API SERVER",   "ip_externa"),
    ("FECHA CREACION",        "fecha_creacion"),
    ("ULTIMA ACTUALIZACION",  "ultima_actualizacion"),
    ("AMBIENTE",              "ambiente"),
    ("CRITICIDAD",            "criticidad"),
    ("PROVEEDOR RESPONSABLE", "proveedor_responsable"),
    ("DESCRIPCION",           "descripcion"),
]

GKE_WIDTHS = {
    "NOMBRE CLUSTER":28, "ESTADO":14, "FUENTE":10, "PROYECTO":32,
    "REGION":24, "ZONA":26, "VPC":28, "SUBRED":28,
    "VERSION KUBERNETES":22, "VERSION NODOS":18,
    "AUTOPILOT":12, "TOTAL NODOS":14, "NODOS MIN":12, "NODOS MAX":12,
    "TIPO MAQUINA NODOS":22, "NODE POOLS":12,
    "RELEASE CHANNEL":18, "WORKLOAD IDENTITY":18,
    "IP CLUSTER (CIDR)":18, "ENDPOINT API SERVER":36,
    "FECHA CREACION":16, "ULTIMA ACTUALIZACION":20,
    "AMBIENTE":14, "CRITICIDAD":12,
    "PROVEEDOR RESPONSABLE":22, "DESCRIPCION":35,
}


def _populate_sheet(ws, resources: list, columns: list, widths: dict) -> None:
    """Llena una hoja generica con los recursos y columnas indicados."""
    ws.freeze_panes = "B2"
    border          = _border()
    headers         = [h for h, _ in columns]
    keys            = [k for _, k in columns]

    for ci, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=ci, value=header)
        cell.font      = Font(bold=True, name="Arial", size=10, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.border    = border
    ws.row_dimensions[1].height = 28

    for ri, resource in enumerate(resources, 2):
        estado    = str(getattr(resource, "estado", "")).upper()
        row_color = ROW_COLORS.get(estado)
        row_fill  = PatternFill("solid", start_color=row_color) if row_color else None

        for ci, key in enumerate(keys, 1):
            # primero buscar en metadata, luego en el resource
            val = resource.metadata.get(key) if resource.metadata else None
            if val is None:
                val = getattr(resource, key, "")
            cell           = ws.cell(row=ri, column=ci, value="" if val is None else val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border    = border
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[ri].height = 16

    for ci, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(header, 15)

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(resources)+1}"


def write(resources: list, output_path: str) -> None:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title       = "INVENTARIO"
    ws.freeze_panes = "B2"

    border = _border()

    # headers
    for ci, header in enumerate(HEADERS, 1):
        cell           = ws.cell(row=1, column=ci, value=header)
        cell.font      = Font(bold=True, name="Arial", size=10, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.border    = border

    ws.row_dimensions[1].height = 28

    # datos
    for ri, resource in enumerate(resources, 2):
        estado    = str(getattr(resource, "estado", "")).upper()
        row_color = ROW_COLORS.get(estado)
        row_fill  = PatternFill("solid", start_color=row_color) if row_color else None

        for ci, key in enumerate(KEYS, 1):
            val  = getattr(resource, key, "")
            cell = ws.cell(row=ri, column=ci, value="" if val is None else val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border    = border
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[ri].height = 16

    # anchos
    for ci, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(header, 15)

    # autofilter
    last_col = get_column_letter(len(HEADERS))
    last_row = len(resources) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # validaciones
    col_amb = get_column_letter(HEADERS.index("AMBIENTE") + 1)
    col_cri = get_column_letter(HEADERS.index("CRITICIDAD") + 1)
    dv_amb  = DataValidation(type="list", formula1='"PRODUCCION,DESARROLLO"', allow_blank=True)
    dv_cri  = DataValidation(type="list", formula1='"ALTA,MEDIA,BAJA"',       allow_blank=True)
    dv_amb.sqref = f"{col_amb}2:{col_amb}{last_row}"
    dv_cri.sqref = f"{col_cri}2:{col_cri}{last_row}"
    ws.add_data_validation(dv_amb)
    ws.add_data_validation(dv_cri)

    wb.save(output_path)
    logger.debug(f"[Excel] Guardado: {output_path} ({len(resources)} recursos)")


def write_all(resources_by_type: dict, output_path: str) -> None:
    """
    Escribe multiples hojas en el Excel segun tipo de recurso.
    El orden y configuracion de cada hoja viene del SHEET_REGISTRY.
    Cloud SQL se fusiona en la hoja MAQUINAS VIRTUALES.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Fusionar Cloud SQL dentro de VIRTUAL MACHINE
    combinado = dict(resources_by_type)
    vm_y_sql  = (
        combinado.get("VIRTUAL MACHINE", []) +
        combinado.get("CLOUD SQL", [])
    )
    vm_y_sql.sort(key=lambda r: (r.tipo_recurso, r.proyecto, r.nombre))
    combinado["VIRTUAL MACHINE"] = vm_y_sql
    combinado.pop("CLOUD SQL", None)

    wb      = Workbook()
    primera = True

    for tipo_recurso, (sheet_title, columns, widths) in SHEET_REGISTRY.items():
        recursos = combinado.get(tipo_recurso, [])
        if not recursos:
            continue

        if primera:
            ws        = wb.active
            ws.title  = sheet_title
            primera   = False
        else:
            ws = wb.create_sheet(sheet_title)

        if tipo_recurso == "VIRTUAL MACHINE":
            _populate_vm_sheet(ws, recursos)
        else:
            _populate_sheet(ws, recursos, columns, widths)

    if primera:
        wb.active.title = "INVENTARIO"

    wb.save(output_path)
    total = sum(len(v) for v in resources_by_type.values())
    hojas = len([t for t in SHEET_REGISTRY if combinado.get(t)])
    logger.debug(f"[Excel] Guardado: {output_path} ({total} recursos, {hojas} hojas)")


def _populate_vm_sheet(ws, resources: list) -> None:
    """Llena la hoja de VMs con el formato estandar."""
    ws.freeze_panes = "B2"
    border = _border()

    for ci, header in enumerate(HEADERS, 1):
        cell           = ws.cell(row=1, column=ci, value=header)
        cell.font      = Font(bold=True, name="Arial", size=10, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.border    = border
    ws.row_dimensions[1].height = 28

    for ri, resource in enumerate(resources, 2):
        estado    = str(getattr(resource, "estado", "")).upper()
        row_color = ROW_COLORS.get(estado)
        row_fill  = PatternFill("solid", start_color=row_color) if row_color else None
        for ci, key in enumerate(KEYS, 1):
            val = resource.metadata.get(key) if resource.metadata else None
            if val is None:
                val = getattr(resource, key, "")
            cell           = ws.cell(row=ri, column=ci, value="" if val is None else val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border    = border
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[ri].height = 16

    for ci, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(header, 15)

    last_col = get_column_letter(len(HEADERS))
    last_row = len(resources) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    col_amb = get_column_letter(HEADERS.index("AMBIENTE") + 1)
    col_cri = get_column_letter(HEADERS.index("CRITICIDAD") + 1)
    dv_amb  = DataValidation(type="list", formula1='"PRODUCCION,DESARROLLO"', allow_blank=True)
    dv_cri  = DataValidation(type="list", formula1='"ALTA,MEDIA,BAJA"',       allow_blank=True)
    dv_amb.sqref = f"{col_amb}2:{col_amb}{last_row}"
    dv_cri.sqref = f"{col_cri}2:{col_cri}{last_row}"
    ws.add_data_validation(dv_amb)
    ws.add_data_validation(dv_cri)


# ── CLOUD SQL ─────────────────────────────────────────────────────────────────
SQL_COLUMNS = [
    ("NOMBRE INSTANCIA",      "nombre"),
    ("ESTADO",                "estado"),
    ("FUENTE",                "fuente"),
    ("PROYECTO",              "proyecto"),
    ("REGION",                "region"),
    ("ZONA",                  "zona"),
    ("MOTOR",                 "motor"),
    ("VERSION",               "version"),
    ("TIER",                  "tier"),
    ("DISCO (GB)",            "disco_gb"),
    ("TIPO DISCO",            "tipo_disco"),
    ("ALTA DISPONIBILIDAD",   "alta_disponibilidad"),
    ("REPLICA LECTURA",       "replica_lectura"),
    ("BACKUP AUTOMATICO",     "backup_automatico"),
    ("VENTANA BACKUP",        "ventana_backup"),
    ("IP PRIVADA",            "ip_privada"),
    ("IP PUBLICA",            "ip_publica"),
    ("FECHA CREACION",        "fecha_creacion"),
    ("ULTIMA ACTUALIZACION",  "ultima_actualizacion"),
    ("AMBIENTE",              "ambiente"),
    ("CRITICIDAD",            "criticidad"),
    ("PROVEEDOR RESPONSABLE", "proveedor_responsable"),
    ("DESCRIPCION",           "descripcion"),
]

SQL_WIDTHS = {
    "NOMBRE INSTANCIA":28, "ESTADO":14, "FUENTE":10, "PROYECTO":32,
    "REGION":24, "ZONA":26,
    "MOTOR":14, "VERSION":26, "TIER":22,
    "DISCO (GB)":11, "TIPO DISCO":12,
    "ALTA DISPONIBILIDAD":20, "REPLICA LECTURA":16,
    "BACKUP AUTOMATICO":18, "VENTANA BACKUP":16,
    "IP PRIVADA":18, "IP PUBLICA":14,
    "FECHA CREACION":16, "ULTIMA ACTUALIZACION":20,
    "AMBIENTE":14, "CRITICIDAD":12,
    "PROVEEDOR RESPONSABLE":22, "DESCRIPCION":35,
}

# ── CLOUD STORAGE ─────────────────────────────────────────────────────────────
STORAGE_COLUMNS = [
    ("NOMBRE BUCKET",         "nombre"),
    ("ESTADO",                "estado"),
    ("FUENTE",                "fuente"),
    ("PROYECTO",              "proyecto"),
    ("REGION",                "region"),
    ("TIPO UBICACION",        "location_type"),
    ("CLASE ALMACENAMIENTO",  "clase_almacenamiento"),
    ("ACCESO PUBLICO",        "acceso_publico"),
    ("VERSIONADO",            "versionado"),
    ("ENCRIPTACION",          "encriptacion"),
    ("RETENTION POLICY",      "retention_policy"),
    ("LIFECYCLE RULES",       "lifecycle_rules"),
    ("FECHA CREACION",        "fecha_creacion"),
    ("ULTIMA ACTUALIZACION",  "ultima_actualizacion"),
    ("AMBIENTE",              "ambiente"),
    ("CRITICIDAD",            "criticidad"),
    ("PROVEEDOR RESPONSABLE", "proveedor_responsable"),
    ("DESCRIPCION",           "descripcion"),
]

STORAGE_WIDTHS = {
    "NOMBRE BUCKET":35, "ESTADO":12, "FUENTE":10, "PROYECTO":32,
    "REGION":24, "TIPO UBICACION":16,
    "CLASE ALMACENAMIENTO":22, "ACCESO PUBLICO":16,
    "VERSIONADO":12, "ENCRIPTACION":26,
    "RETENTION POLICY":18, "LIFECYCLE RULES":16,
    "FECHA CREACION":16, "ULTIMA ACTUALIZACION":20,
    "AMBIENTE":14, "CRITICIDAD":12,
    "PROVEEDOR RESPONSABLE":22, "DESCRIPCION":35,
}

# ── APIS HABILITADAS ──────────────────────────────────────────────────────────
APIS_COLUMNS = [
    ("API ID",                "api_id"),
    ("NOMBRE LEGIBLE",        "nombre_legible"),
    ("ESTADO",                "estado"),
    ("FUENTE",                "fuente"),
    ("PROYECTO",              "proyecto"),
    ("ULTIMA ACTUALIZACION",  "ultima_actualizacion"),
]

APIS_WIDTHS = {
    "API ID":40, "NOMBRE LEGIBLE":45, "ESTADO":14,
    "FUENTE":10, "PROYECTO":32, "ULTIMA ACTUALIZACION":20,
}

# ── CLOUD RUN ─────────────────────────────────────────────────────────────────
CLOUDRUN_COLUMNS = [
    ("NOMBRE SERVICIO",       "nombre"),
    ("ESTADO",                "estado"),
    ("FUENTE",                "fuente"),
    ("PROYECTO",              "proyecto"),
    ("REGION",                "region"),
    ("URL",                   "ip_externa"),
    ("CPU",                   "vcpus"),
    ("MEMORIA (GB)",          "ram_gb"),
    ("MIN INSTANCIAS",        "min_instancias"),
    ("MAX INSTANCIAS",        "max_instancias"),
    ("CONCURRENCIA",          "concurrencia"),
    ("INGRESS",               "ingress"),
    ("ULTIMA REVISION",       "ultima_revision"),
    ("FECHA CREACION",        "fecha_creacion"),
    ("ULTIMA ACTUALIZACION",  "ultima_actualizacion"),
    ("AMBIENTE",              "ambiente"),
    ("CRITICIDAD",            "criticidad"),
    ("PROVEEDOR RESPONSABLE", "proveedor_responsable"),
    ("DESCRIPCION",           "descripcion"),
]

CLOUDRUN_WIDTHS = {
    "NOMBRE SERVICIO":30, "ESTADO":14, "FUENTE":10, "PROYECTO":32,
    "REGION":24, "URL":50,
    "CPU":8, "MEMORIA (GB)":14,
    "MIN INSTANCIAS":16, "MAX INSTANCIAS":16,
    "CONCURRENCIA":14, "INGRESS":14,
    "ULTIMA REVISION":36,
    "FECHA CREACION":16, "ULTIMA ACTUALIZACION":20,
    "AMBIENTE":14, "CRITICIDAD":12,
    "PROVEEDOR RESPONSABLE":22, "DESCRIPCION":35,
}

# ── REGISTRO DE HOJAS ─────────────────────────────────────────────────────────
# CLOUD SQL se fusiona en MAQUINAS VIRTUALES — no tiene hoja propia
SHEET_REGISTRY = {
    "VIRTUAL MACHINE": ("MAQUINAS VIRTUALES", None,               None),
    "GKE CLUSTER":     ("GKE",                GKE_COLUMNS,        GKE_WIDTHS),
    "CLOUD STORAGE":   ("ALMACENAMIENTO",     STORAGE_COLUMNS,    STORAGE_WIDTHS),
    "API HABILITADA":  ("SERVICIOS",          APIS_COLUMNS,       APIS_WIDTHS),
    "CLOUD RUN":       ("CLOUD RUN",          CLOUDRUN_COLUMNS,   CLOUDRUN_WIDTHS),
}
