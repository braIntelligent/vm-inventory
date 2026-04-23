"""
merger.py
Logica idempotente de merge entre el estado actual (API) y el Excel existente.
- Recursos nuevos     → se agregan con campos manuales vacios
- Recursos existentes → se actualizan solo los campos automaticos
- Recursos eliminados → se marcan como ELIMINADA solo si su fuente estuvo disponible
"""

import logging
import os

import openpyxl

from core.models import Resource

logger = logging.getLogger(__name__)

MANUAL_COLS = [
    "ambiente", "criticidad", "categoria",
    "proveedor_responsable", "pam",
    "version_cylance", "version_focus", "grupo_tenable", "descripcion",
]

HEADER_MAP = {
    "NOMBRE":                "nombre",
    "NOMBRE VM":             "nombre",   # backwards compat con Excel anterior
    "ESTADO":                "estado",
    "TIPO RECURSO":          "tipo_recurso",
    "FUENTE":                "fuente",
    "PROYECTO":              "proyecto",
    "REGION":                "region",
    "ZONA":                  "zona",
    "VPC":                   "vpc",
    "SUBRED":                "subred",
    "TIPO MAQUINA":          "tipo_maquina",
    "VCPUS":                 "vcpus",
    "RAM (GB)":              "ram_gb",
    "DISCO (GB)":            "disco_gb",
    "TIPO DISCO":            "tipo_disco",
    "SISTEMA OPERATIVO":     "sistema_operativo",
    "IP INTERNA":            "ip_interna",
    "IP EXTERNA":            "ip_externa",
    "FECHA CREACION":        "fecha_creacion",
    "ULTIMO ENCENDIDO":      "ultimo_encendido",
    "ULTIMO APAGADO":        "ultimo_apagado",
    "PROGRAMA DE ENCENDIDO": "programa_encendido",
    "ULTIMA ACTUALIZACION":  "ultima_actualizacion",
    "AMBIENTE":              "ambiente",
    "CRITICIDAD":            "criticidad",
    "CATEGORIA":             "categoria",
    "PROVEEDOR RESPONSABLE": "proveedor_responsable",
    "PAM":                   "pam",
    "VERSION CYLANCE":       "version_cylance",
    "VERSION FOCUS":         "version_focus",
    "GRUPO TENABLE":         "grupo_tenable",
    "DESCRIPCION":           "descripcion",
}


def _load_existing(path: str) -> dict:
    if not os.path.exists(path):
        logger.debug("[Merger] Sin Excel previo — primera ejecucion desde cero.")
        return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        logger.warning(f"[Merger] Excel existente no valido, arrancando desde cero: {e}")
        return {}

    ws          = wb.active
    headers_row = [cell.value for cell in ws[1]]
    col_keys    = [HEADER_MAP.get(str(h).strip(), "") for h in headers_row]

    existing = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for key, val in zip(col_keys, row):
            if key:
                record[key] = val if val is not None else ""
        name    = str(record.get("nombre", "")).strip()
        project = str(record.get("proyecto", "")).strip()
        key     = f"{name}::{project}"
        if name:
            try:
                existing[key] = Resource.from_dict(record)
            except Exception as e:
                logger.debug(f"  Error reconstruyendo Resource {key}: {e}")

    logger.debug(f"[Merger] {len(existing)} recursos en Excel existente.")
    return existing


def _make_key(r) -> str:
    """
    Clave unica por recurso: nombre + proyecto.
    Evita colisiones en recursos como APIs que tienen el mismo nombre
    en distintos proyectos.
    """
    return f"{r.nombre}::{r.proyecto}"


def merge(new_resources: list, output_path: str) -> list:
    existing        = _load_existing(output_path)
    new_map         = {_make_key(r): r for r in new_resources}
    fuentes_activas = {r.fuente for r in new_resources}
    now             = new_resources[0].ultima_actualizacion if new_resources else ""

    final = []

    for key, new_r in new_map.items():
        if key in existing:
            old_r = existing[key]
            for col in MANUAL_COLS:
                val = getattr(old_r, col, "")
                if val:
                    setattr(new_r, col, val)
            logger.debug(f"  [UPDATE]   {key}")
        else:
            logger.debug(f"  [NUEVA]    {key}")
        final.append(new_r)

    for key, old_r in existing.items():
        if key not in new_map:
            if old_r.fuente in fuentes_activas:
                old_r.estado               = "ELIMINADA"
                old_r.ultima_actualizacion = now
                logger.debug(f"  [ELIMINADA] {key}")
            else:
                logger.debug(f"  [CONSERVADA] {key}")
            final.append(old_r)

    final.sort(key=lambda r: (r.tipo_recurso, r.proyecto, r.nombre))
    logger.debug(f"[Merger] Total final: {len(final)} recursos.")
    return final
